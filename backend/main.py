import openpyxl
import time
import log_modules
import generate_barcode

# File path for the Excel workbook
FILE_PATH = 'Filament-Logs\\filament_inventory.xlsx'

# Initialize workbook and sheet
try:
    workbook = openpyxl.load_workbook(FILE_PATH)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Timestamp', 'Barcode', 'Brand', 'Color', 'Material', 'Attribute 1', 'Attribute 2', 'Filament Amount (g)', 'Location', 'Roll Weight (g)', 'Times Logged Out'])  # Add headers

def log_filament_data():
    """
    Main loop to log filament data, including barcode, weight, and decoded filament name.
    Saves the data to an Excel file and handles user interruptions gracefully.
    """
    print("Press Ctrl+C to exit the program.")

    while True:
        try:
            # Gather data
            barcode = log_modules.get_barcode()
            brand, color, material, attribute_1, attribute_2, location = log_modules.decode_barcode(barcode)

            # Display log info
            print('Logging weight for filament: {} {} {}{}{}'.format(
                brand,
                color,
                material,
                f' {attribute_1}' if attribute_1 else '',
                f' {attribute_2}' if attribute_2 else ''
            ))

            roll_weight = log_modules.get_roll_weight(barcode, sheet)
            filament_amount = log_modules.get_current_weight(roll_weight)

            # Find and update the existing row
            for row in sheet.iter_rows(min_row=2):
                if row[1].value == barcode:  # Barcode is in the second column
                    row[7].value = filament_amount          # Update filament amount (8th column)
                    row[10].value = int(row[10].value) + 1  # Add 1 to the amount of times logged out (11th column)
                    break
            workbook.save(FILE_PATH)

            print(f'''Updated: 
                    Barcode: {barcode}, 
                    Brand: {brand}, 
                    Color: {color}, 
                    Material: {material},''') 
            
            if attribute_1 != '':
                print(f'Attribute 1: {attribute_1},')
            if attribute_2 != '':
                print(f'Attribute 2: {attribute_2},')

            print(f'New Filament Amount: {filament_amount},')
            print(f'Location: {location}')

        except KeyboardInterrupt:
            print('\nExiting program. Goodbye!')
            break
        except Exception as e:
            print(f"An error occurred: {e}")

def log_full_filament_data():
    """
    Main loop to generate new barcodes and log new rolls of filament.
    Saves the data to an Excel file and handles user interruptions gracefully.
    """
    print("Press Ctrl+C to exit the program.")
    
    while True:
        try:
            brand = input('Enter the brand: ')
            color = input('Enter the color: ')
            material = input('Enter the material: ')
            attribute_1 = input('Enter the first attribute: ')
            attribute_2 = input('Enter the second attribute: ')
            location = input('Enter the location of the filament: ')

            barcode = generate_barcode.generate_filament_barcode(brand, color, material, attribute_1, attribute_2, location, sheet)
            brand, color, material, attribute_1, attribute_2, location = log_modules.decode_barcode(barcode)
            
            print('New barcode for {} {} {}{}{} in {} is {}'.format(
                brand,
                color,
                material,
                f' {attribute_1}' if attribute_1 else '',
                f' {attribute_2}' if attribute_2 else '',
                location,
                barcode
            ))
            
            # Append data to the sheet
            timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
            roll_weight, filament_amount = log_modules.get_starting_weight()

            sheet.append([timestamp, barcode, brand, color, material, attribute_1, attribute_2, filament_amount, location, roll_weight, '0'])
            workbook.save(FILE_PATH)

            print(f'''Logged: 
                    Timestamp: {timestamp}, 
                    Barcode: {barcode}, 
                    Brand: {brand}, 
                    Color: {color}, 
                    Material: {material},''') 
            
            if attribute_1 != '':
                print(f'Attribute 1: {attribute_1},')
            if attribute_2 != '':
                print(f'Attribute 2: {attribute_2},')
                
            print(f'''Filament Amount: {filament_amount}, 
                    Location: {location}, 
                    Roll Weight: {roll_weight}''')
            
        except KeyboardInterrupt:
            print('\nExiting program. Goodbye!')
            break
        except ValueError as e:
            print(e)

if __name__ == '__main__':
    
    user = None

    if user == 'Admin':
        log_full_filament_data()
    else:
        log_filament_data()