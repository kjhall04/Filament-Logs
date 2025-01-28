import openpyxl
import json
import difflib

# File path for the Excel workbook
FILE_PATH = 'filament_inventory.xlsx'

def generate_filament_barcode(brand: str, color: str, material: str, sheet) -> str:
    """
    Generate a barcode for a new filament roll based on the provided information, with fuzzy matching.
    
    Args:
        brand (str): The brand name (e.g., "Brand A").
        color (str): The color name (e.g., "White").
        material (str): The material name (e.g., "PLA").
        sheet: The Excel sheet object to read existing barcodes from.
    
    Returns:
        str: A 13-digit numeric barcode.
    """
    brand_mapping = load_json('brand_mapping.json')
    color_mapping = load_json('color_mapping.json')
    material_mapping = load_json('material_mapping.json')

    # Get closest matches for material, color, and brand
    brand_code = get_closest_match(brand, brand_mapping, "Invalid Brand")
    color_code = get_closest_match(color, color_mapping, "Invalid Color")
    material_code = get_closest_match(material, material_mapping, "Invalid Material")

    # Validate that the closest matches are found
    if brand_code == "Invalid Brand":
        raise ValueError(f"Invalid brand: {brand}. Please use a valid brand name.")
    if color_code == "Invalid Color":
        raise ValueError(f"Invalid color: {color}. Please use a valid color name.")
    if material_code == "Invalid Material":
        raise ValueError(f"Invalid material: {material}. Please use a valid material name.")
    
    # Extract existing barcodes from the sheet
    barcodes = [row[1].value for row in sheet.iter_rows(min_row=2) if row[1].value]

    # Extract unique IDs from existing barcode
    unique_ids = [int(barcode[-6:]) for barcode in barcodes if barcode.isdigit() and len(barcode) == 13]

    # Determine the next unique ID
    next_unique_id = max(unique_ids, default = 0) + 1

    # Format unique ID as zero-padded string (up to 6 digits)
    unique_id_str = f"{next_unique_id:06}"

    # Construct the barcode
    barcode = f"{brand_code}{color_code}{material_code}{unique_id_str}"
    return barcode

def get_closest_match(name, mapping, default):
    """
    Find the closest match for a name in a given mapping using fuzzy matching.

    Args:
        name (str): The name to match.
        mapping (dict): The mapping dictionary to search in.
        default (str): The default value if no close match is found.

    Returns:
        str: The matched code or the default.
    """
    keys = list(mapping.keys())
    values = list(mapping.values())
    matches = difflib.get_close_matches(name, values, n=1, cutoff=0.6)
    if matches:
        matched_value = matches[0]
        # Return the corresponding key for the matched value
        return next((key for key, value in mapping.items() if value == matched_value), default)
    return default

def load_json(filename):
    """
    Load a JSON file and return its content.

    Args:
        filename (str): The name of the JSON file.

    Returns:
        dict: The content of the JSON file.
    """
    with open(filename, 'r') as file:
        return json.load(file)

if __name__ == '__main__':
    # Open the Excel workbook and sheet
    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Timestamp', 'Barcode', 'Brand', 'Color', 'Material', 'Weight (g)'])  # Add headers

    while True:
        brand = input('Enter the brand: ')  # Brand first
        color = input('Enter the color: ')
        material = input('Enter the material: ')  # Material last

        try:
            barcode = generate_filament_barcode(brand, color, material, sheet)
            print(f'New barcode is {barcode}')
        except ValueError as e:
            print(e)