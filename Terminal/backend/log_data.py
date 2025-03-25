import openpyxl
import time
import data_manipulation
import generate_barcode

# File path for the Excel workbook
FILE_PATH = 'Filament-Logs\\filament_inventory.xlsx'

# Empty threshold for filament detection
EMPTY_THRESHOLD = 5

# Initialize workbook and sheet
try:
    workbook = openpyxl.load_workbook(FILE_PATH)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Timestamp', 'Barcode', 'Brand', 'Color', 'Material', 'Attribute 1', 'Attribute 2', 
                  'Filament Amount (g)', 'Location', 'Roll Weight (g)', 'Times Logged Out', 'Is Empty']) # Add headers

def log_filament_data():
    """ Logs filament data including barcode, weight, and updates Excel. """
    print("Press Ctrl+C to exit the program.\n\nPlace filament on the scale now.\n")

    while True:
        try:
            # Gather data
            barcode = data_manipulation.get_barcode()
            brand, color, material, attr1, attr2, location = data_manipulation.decode_barcode(barcode)

            roll_weight = data_manipulation.get_roll_weight(barcode, sheet)
            filament_amount = data_manipulation.get_current_weight(roll_weight)
            timestamp = time.strftime('%Y-%m-%d %H:%M:%S')

            # Display log info
            attributes = " ".join(filter(None, [attr1, attr2])) # Removes empty attributes
            print(f'Logging weight for filament: {brand} {color} {material} {attributes}')

            # Find and update the existing row
            for row in sheet.iter_rows(min_row=2):
                if row[1].value == barcode:  # Barcode is in the second column
                    row[0].value = timestamp  # Update timestamp
                    row[7].value = filament_amount  # Update filament amount (8th column)
                    row[10].value = int(row[10].value) + 1  # Add 1 to the amount of times logged out (11th column)

                    # Check if roll is empty
                    if int(filament_amount) <= EMPTY_THRESHOLD:
                        user_confirm = input(f"Filament weight is {filament_amount}g. Mark as empty? (y/n): ").strip().lower()
                        if user_confirm == 'y':
                            row[11].value = 'True'  # Mark as empty in Excel

                    break  # Exit loop after finding the barcode
            
            # Save Changes
            workbook.save(FILE_PATH)

            print(f'''Updated: 
                    Barcode: {barcode}, 
                    Brand: {brand}, 
                    Color: {color}, 
                    Material: {material},
                    Attributes: {attributes},
                    New Filament Amount: {filament_amount},
                    Location: {location}''') 
            
        except KeyboardInterrupt:
            print('\nExiting program. Goodbye!')
            break
        except Exception as e:
            print(f"An error occurred: {e}")

def log_full_filament_data():
    """ Generates new barcodes and logs full rolls of filament. """
    print("Press Ctrl+C to exit the program.\n\nPlace filament on the scale now.\n")
    
    while True:
        try:
            # Gather user input
            brand = input('Enter the brand: ')
            color = input('Enter the color: ')
            material = input('Enter the material: ')
            attr1 = input('Enter the first attribute: ')
            attr2 = input('Enter the second attribute: ')
            location = input('Enter the location of the filament: ')

            # Generate barcode
            barcode = generate_barcode.generate_filament_barcode(brand, color, material, attr1, attr2, location, sheet)
            brand, color, material, attr1, attr2, location = data_manipulation.decode_barcode(barcode)

            attributes = " ".join(filter(None, [attr1, attr2]))
            print(f'New barcode for {brand} {color} {material} {attributes} in {location}: {barcode}')

            # Get filament weight details
            timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
            roll_weight, filament_amount = data_manipulation.get_starting_weight()

            # Append data to the sheet
            sheet.append([timestamp, barcode, brand, color, material, attr1, attr2, filament_amount, 
                          location, roll_weight, '0', 'False'])
            workbook.save(FILE_PATH)

            print(f'''Logged: 
                    Timestamp: {timestamp}, 
                    Barcode: {barcode}, 
                    Brand: {brand}, 
                    Color: {color}, 
                    Material: {material},
                    Attributes: {attributes},
                    Filament Amount: {filament_amount}, 
                    Location: {location}, 
                    Roll Weight: {roll_weight}''') 
            
        except KeyboardInterrupt:
            print('\nExiting program. Goodbye!')
            break
        except ValueError as e:
            print(e)

if __name__ == '__main__':
    user = None

    if user == 'Admin':
        log_full_filament_data()
    else:
        log_filament_data()