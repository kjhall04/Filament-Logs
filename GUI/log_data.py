import openpyxl
import time
import data_manipulation as dm
import generate_barcode as gb

FILE_PATH = 'Filament-Logs\\filament_inventory.xlsx'
EMPTY_THRESHOLD = 5

def load_workbook():
    """Load the existing workbook or create a new one."""
    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([
            'Timestamp', 'Barcode', 'Brand', 'Color', 'Material', 
            'Attribute 1', 'Attribute 2', 'Filament Amount (g)', 
            'Location', 'Roll Weight (g)', 'Times Logged Out', 'Is Empty'
        ])
        workbook.save(FILE_PATH)  # Save the new workbook
    return workbook
    
workbook = load_workbook()
sheet = workbook.active

def log_filament_data(barcode, filament_amount):
    """ Logs filament data including barcode, weight, and updates Excel. """
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')

    try:
        # Find and update the existing row
        for row in sheet.iter_rows(min_row=2):
            if row[1].value == barcode:  # Barcode is in the second column
                row[0].value = timestamp  # Update timestamp
                row[7].value = filament_amount  # Update filament amount (8th column)
                row[10].value = int(row[10].value) + 1  # Add 1 to the amount of times logged out (11th column)

                # Check if roll is empty
                if int(filament_amount) <= EMPTY_THRESHOLD:
                    row[11].value = 'True'  # Mark as empty in Excel
                break  # Exit loop after finding the barcode
            
            # Save Changes
            workbook.save(FILE_PATH)

    except Exception as e:
        return f"An error occurred: {e}"
    else:
        return None

def log_full_filament_data(barcode):
    """ Generates new barcodes and logs full rolls of filament. """
    try:
        # Generate barcode
        brand, color, material, attr1, attr2, location = dm.decode_barcode(barcode)

        # Get filament weight details
        timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
        roll_weight, filament_amount = dm.get_starting_weight()

        # Append data to the sheet
        sheet.append([timestamp, barcode, brand, color, material, attr1, attr2, filament_amount, 
                        location, roll_weight, '0', 'False'])
        workbook.save(FILE_PATH)

    except ValueError as e:
        return e
    else:
        return None

if __name__ == '__main__':
    user = 'Admin'

    if user == 'Admin':
        log_full_filament_data()
    else:
        log_filament_data()