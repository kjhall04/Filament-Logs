from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
import os
from datetime import datetime
from dotenv import load_dotenv

from backend import spreadsheet_stats
from backend import log_data
from backend import generate_barcode
from backend import data_manipulation

load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY")

EXCEL_PATH = r"C:\Users\LichKing\Desktop\Programming\Filament-Logs\filament_inventory.xlsx"

def get_sheet():
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Barcode', 'Brand', 'Color', 'Material', 'Attribute 1', 'Attribute 2', 
                   'Filament Amount (g)', 'Location', 'Roll Weight (g)', 'Times Logged Out', 'Is Empty', 
                   'Is Favorite'])
        wb.save(EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    return wb, wb.active

@app.route("/")
def index():
    wb, sheet = get_sheet()
    filaments = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    # Sort by timestamp (first column), descending
    filaments.sort(
        key=lambda x: datetime.strptime(str(x[0]), "%Y-%m-%d %H:%M:%S") if x[0] else datetime.min,
        reverse=True
    )
    # Build favorites list
    favorite_barcodes = [f[1] for f in filaments if len(f) > 12 and str(f[12]).lower() == "true"]
    return render_template(
        "index.html",
        filaments=filaments,
        total=len(filaments),
        favorite_barcodes=favorite_barcodes
    )

@app.route("/popular")
def popular_filaments():
    wb, sheet = get_sheet()
    popular = spreadsheet_stats.get_most_popular_filaments()
    return render_template("popular.html", filaments=popular)

@app.route("/low_empty")
def low_empty_filaments():
    wb, sheet = get_sheet()
    low_empty = spreadsheet_stats.get_low_or_empty_filaments()
    return render_template("low_empty.html", filaments=low_empty)

@app.route("/empty_rolls")
def empty_rolls():
    wb, sheet = get_sheet()
    empty = spreadsheet_stats.get_empty_rolls()
    return render_template("empty_rolls.html", rolls=empty)

@app.route("/log", methods=["GET", "POST"])
def log_filament():
    if request.method == "POST":
        barcode = request.form["barcode"]
        weight = request.form["weight"]
        roll_weight_val = data_manipulation.get_roll_weight(barcode, get_sheet()[1])
        filament_amount = int(weight) - int(roll_weight_val)
        result = log_data.log_filament_data_web(barcode, filament_amount, roll_weight_val)
        flash("Filament usage logged successfully!", "success")
        return redirect(url_for("index")) 
    return render_template("log.html")

@app.route("/new_roll", methods=["GET", "POST"])
def new_roll():
    wb, sheet = get_sheet()
    # Step 1: Enter filament info and generate barcode
    if request.method == "POST" and "step" not in request.form:
        brand = request.form.get("brand", "")
        color = request.form.get("color", "")
        material = request.form.get("material", "")
        attr1 = request.form.get("attribute_1", "")
        attr2 = request.form.get("attribute_2", "")
        location = request.form.get("location", "")

        # Generate barcode
        barcode = generate_barcode.generate_filament_barcode(
            brand, color, material, attr1, attr2, location, sheet
        )

        # Read weight from scale using data_manipulation
        scale_weight = None
        try:
            scale_weight, _ = data_manipulation.get_starting_weight()
        except Exception:
            scale_weight = ""

        # Show barcode and prompt for current weight (pre-fill from scale)
        return render_template(
            "new_roll.html",
            barcode=barcode,
            brand=brand,
            color=color,
            material=material,
            attribute_1=attr1,
            attribute_2=attr2,
            location=location,
            scale_weight=scale_weight,
            step="weight"
        )

    # Step 2: Enter current weight after barcode is generated
    if request.method == "POST" and request.form.get("step") == "weight":
        brand = request.form.get("brand", "")
        color = request.form.get("color", "")
        material = request.form.get("material", "")
        attr1 = request.form.get("attribute_1", "")
        attr2 = request.form.get("attribute_2", "")
        location = request.form.get("location", "")
        barcode = request.form.get("barcode", "")
        starting_weight = int(request.form.get("weight", ""))

        # Calculate roll weight for new roll
        FILAMENT_AMOUNT = data_manipulation.FILAMENT_AMOUNT  # e.g., 1000
        roll_weight = starting_weight - FILAMENT_AMOUNT
        filament_amount = starting_weight - roll_weight  # Should be FILAMENT_AMOUNT

        # Log to spreadsheet (include Is Favorite field as last column)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.append([
            timestamp, barcode, brand, color, material, attr1, attr2,
            filament_amount, location, roll_weight, '0', 'False', 'False'
        ])
        wb.save(EXCEL_PATH)
        flash(f"New filament roll added! Barcode: {barcode}, Filament Amount: {filament_amount}", "success")
        return redirect(url_for("index"))

    # Initial GET: show info form
    return render_template("new_roll.html", step="info")

@app.route("/toggle_favorite", methods=["POST"])
def toggle_favorite():
    barcode = request.json.get("barcode")
    wb, sheet = get_sheet()
    for row in sheet.iter_rows(min_row=2):
        if str(row[1].value) == barcode:
            current = str(row[12].value).lower() if row[12].value else "false"
            row[12].value = "False" if current == "true" else "True"
            break
    wb.save(EXCEL_PATH)
    return '', 204

@app.route("/favorites")
def favorites():
    wb, sheet = get_sheet()
    rows = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Ensure threshold is numeric
    try:
        threshold = float(getattr(log_data, "EMPTY_THRESHOLD", 250))
    except Exception:
        threshold = 250.0

    # Helpers
    def key_norm(val):
        # normalization for grouping keys: lower-case + trim
        if val is None:
            return ""
        return str(val).strip().lower()

    def display_norm(val):
        # trimmed value for display (preserve case)
        if val is None:
            return ""
        return str(val).strip()

    import re
    number_re = re.compile(r"[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?|[-+]?\d*\.?\d+")
    def parse_number(v):
        """Try to extract a numeric value from v. Returns float or None."""
        if v is None:
            return None
        # If it's already a number type, coerce to float
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s == "":
            return None
        # remove common unit words and trailing g/G etc
        # Use regex to extract the first numeric token (handles 1,000 and decimals)
        m = number_re.search(s.replace(',', ''))
        if not m:
            return None
        try:
            return float(m.group())
        except Exception:
            return None

    # Build counts for groups (brand, color, attr1, attr2)
    counts = {}
    for r in rows:
        brand_k = key_norm(r[2]) if len(r) > 2 else ""
        color_k = key_norm(r[3]) if len(r) > 3 else ""
        attr1_k = key_norm(r[5]) if len(r) > 5 else ""
        attr2_k = key_norm(r[6]) if len(r) > 6 else ""
        key = (brand_k, color_k, attr1_k, attr2_k)
        entry = counts.setdefault(key, {"total": 0, "low": 0})
        entry["total"] += 1

        # Determine empty flag
        is_empty = False
        if len(r) > 11 and r[11] is not None:
            is_empty = str(r[11]).strip().lower() == "true"

        # Parse filament amount safely
        filament_amount = None
        if len(r) > 7:
            filament_amount = parse_number(r[7])

        # Count as low if empty OR amount < threshold
        if is_empty or (filament_amount is not None and filament_amount < threshold):
            entry["low"] += 1

    # Collect unique favorite groups and attach counts (grouping excludes material)
    unique_favorites = {}
    for f in rows:
        if len(f) > 12 and f[12] is not None and str(f[12]).strip().lower() == "true":
            brand_disp = display_norm(f[2]) if len(f) > 2 else ""
            color_disp = display_norm(f[3]) if len(f) > 3 else ""
            material_disp = display_norm(f[4]) if len(f) > 4 else ""
            attr1_disp = display_norm(f[5]) if len(f) > 5 else ""
            attr2_disp = display_norm(f[6]) if len(f) > 6 else ""

            fav_key = (brand_disp.lower(), color_disp.lower(), material_disp.lower(), attr1_disp.lower(), attr2_disp.lower())
            if fav_key in unique_favorites:
                continue

            # Build Amazon search URL
            query = " ".join([brand_disp, color_disp, material_disp, attr1_disp, attr2_disp, "filament"]).strip()
            amazon_url = "https://www.amazon.com/s?k=" + "+".join(query.split()) if query else ""

            # Use same grouping key used when computing counts (brand,color,attr1,attr2)
            counts_key = (brand_disp.lower(), color_disp.lower(), attr1_disp.lower(), attr2_disp.lower())
            c = counts.get(counts_key, {"total": 0, "low": 0})

            unique_favorites[fav_key] = {
                "brand": f[2],
                "color": f[3],
                "material": f[4],
                "attribute_1": f[5],
                "attribute_2": f[6],
                "amazon_url": amazon_url,
                "total_count": c["total"],
                "low_count": c["low"]
            }

    return render_template("favorites.html", favorites=list(unique_favorites.values()))

if __name__ == "__main__":
    app.run(debug=True)