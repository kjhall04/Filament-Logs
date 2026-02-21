import openpyxl
from datetime import datetime, timedelta
from backend.config import EXCEL_PATH, LOW_THRESHOLD

def _load_sheet(path: str = EXCEL_PATH):
    """Return the active sheet for the given workbook path."""
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb.active

def _parse_timestamp(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v), fmt)
        except Exception:
            continue
    # last resort: try float/int -> excel serial (not implemented) -> return None
    return None

def get_most_popular_filaments(file_path: str = EXCEL_PATH, top_n: int = 10, weeks: int | None = None):
    """
    Return a list of dicts for the most popular filaments sorted by times_logged_out desc.
    If `weeks` is provided (e.g. weeks=4) only rows with last-logged timestamp within that window
    are considered. Note: spreadsheet only stores the latest timestamp per roll.
    Each dict contains: brand, color, material, attribute_1, attribute_2, times_logged_out, weight, is_favorite
    """
    sheet = _load_sheet(file_path)
    popular_filaments = []
    cutoff = None
    if weeks is not None:
        cutoff = datetime.now() - timedelta(weeks=weeks)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # if a cutoff is set, parse timestamp and skip rows older than cutoff
        if cutoff is not None:
            last_logged = _parse_timestamp(row[0] if len(row) > 0 else None)
            if last_logged is None or last_logged < cutoff:
                continue

        brand = row[2] if len(row) > 2 else None
        color = row[3] if len(row) > 3 else None
        material = row[4] if len(row) > 4 else None
        attribute_1 = row[5] if len(row) > 5 else None
        attribute_2 = row[6] if len(row) > 6 else None
        times_logged_out = 0
        if len(row) > 10 and row[10] is not None:
            try:
                times_logged_out = int(row[10])
            except Exception:
                try:
                    times_logged_out = int(float(row[10]))
                except Exception:
                    times_logged_out = 0
        weight = row[7] if len(row) > 7 else None
        favorite = 'false'
        if len(row) > 12 and row[12] is not None:
            favorite = str(row[12]).lower()

        popular_filaments.append({
            'brand': brand,
            'color': color,
            'material': material,
            'attribute_1': attribute_1,
            'attribute_2': attribute_2,
            'times_logged_out': times_logged_out,
            'weight': weight,
            'is_favorite': favorite
        })

    popular_filaments.sort(key=lambda x: x.get('times_logged_out', 0), reverse=True)
    return popular_filaments[:top_n]

def get_low_or_empty_filaments(file_path: str = EXCEL_PATH, low_threshold: float = LOW_THRESHOLD):
    """
    Return filaments that are marked empty or have filament amount below low_threshold.
    Each dict contains: brand, color, material, attribute_1, attribute_2, weight,
    is_favorite, is_empty, last_logged
    """
    sheet = _load_sheet(file_path)
    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        last_logged = row[0] if len(row) > 0 else None
        brand = row[2] if len(row) > 2 else None
        color = row[3] if len(row) > 3 else None
        material = row[4] if len(row) > 4 else None
        attribute_1 = row[5] if len(row) > 5 else None
        attribute_2 = row[6] if len(row) > 6 else None
        filament_amount = None
        if len(row) > 7 and row[7] is not None:
            try:
                filament_amount = float(row[7])
            except Exception:
                filament_amount = None
        is_empty = 'false'
        if len(row) > 11 and row[11] is not None:
            is_empty = str(row[11]).lower()
        favorite = 'false'
        if len(row) > 12 and row[12] is not None:
            favorite = str(row[12]).lower()

        if is_empty == 'true' or (filament_amount is not None and filament_amount < low_threshold):
            results.append({
                'last_logged': last_logged,
                'brand': brand,
                'color': color,
                'material': material,
                'attribute_1': attribute_1,
                'attribute_2': attribute_2,
                'weight': filament_amount,
                'is_empty': is_empty,
                'is_favorite': favorite
            })

    return results

def get_empty_rolls(file_path: str = EXCEL_PATH):
    """
    Return rows marked as empty. Each dict contains: brand, color, material, attribute_1, attribute_2,
    times_logged_out, last_logged (string or None), is_favorite. Sorted newest last_logged first when possible.
    """
    sheet = _load_sheet(file_path)
    empty_rolls = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        last_logged = row[0] if len(row) > 0 else None
        brand = row[2] if len(row) > 2 else None
        color = row[3] if len(row) > 3 else None
        material = row[4] if len(row) > 4 else None
        attribute_1 = row[5] if len(row) > 5 else None
        attribute_2 = row[6] if len(row) > 6 else None
        times_logged_out = 0
        if len(row) > 10 and row[10] is not None:
            try:
                times_logged_out = int(row[10])
            except Exception:
                try:
                    times_logged_out = int(float(row[10]))
                except Exception:
                    times_logged_out = 0
        is_empty = 'false'
        if len(row) > 11 and row[11] is not None:
            is_empty = str(row[11]).lower()
        favorite = 'false'
        if len(row) > 12 and row[12] is not None:
            favorite = str(row[12]).lower()

        if is_empty == 'true':
            empty_rolls.append({
                'brand': brand,
                'color': color,
                'material': material,
                'attribute_1': attribute_1,
                'attribute_2': attribute_2,
                'times_logged_out': times_logged_out,
                'last_logged': last_logged,
                'is_favorite': favorite
            })

    def _parse_last_logged(v):
        if v is None:
            return datetime.min
        if isinstance(v, datetime):
            return v
        try:
            return datetime.strptime(str(v), "%Y-%m-%d %H:%M:%S")
        except Exception:
            return datetime.min

    empty_rolls.sort(key=lambda x: _parse_last_logged(x.get('last_logged')), reverse=True)
    return empty_rolls
