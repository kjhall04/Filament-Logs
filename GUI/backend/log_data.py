import os
import openpyxl
from datetime import datetime
from backend import generate_barcode
from backend.config import EXCEL_PATH, EMPTY_THRESHOLD

HEADERS = ['Timestamp', 'Barcode', 'Brand', 'Color', 'Material', 'Attribute 1', 'Attribute 2',
           'Filament Amount (g)', 'Location', 'Roll Weight (g)', 'Times Logged Out', 'Is Empty',
           'Is Favorite']

def get_workbook_and_sheet():
    """Return (workbook, sheet), creating the file with headers if needed."""
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(HEADERS)
        wb.save(EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    return wb, wb.active

def log_filament_data_web(barcode, filament_amount, roll_weight=None):
    """
    Update an existing roll (identified by barcode) with a new filament_amount.
    Increments Times Logged Out and sets Is Empty based on EMPTY_THRESHOLD.
    Returns True if an existing row was updated, False otherwise.
    """
    wb, sheet = get_workbook_and_sheet()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    updated = False

    for row in sheet.iter_rows(min_row=2):
        cell_barcode = row[1].value
        if cell_barcode is not None and str(cell_barcode).strip() == str(barcode).strip():
            row[0].value = timestamp
            # store filament amount
            try:
                row[7].value = int(round(float(filament_amount)))
            except Exception:
                row[7].value = filament_amount
            # increment times logged out
            try:
                row[10].value = int(row[10].value) + 1 if row[10].value is not None else 1
            except Exception:
                row[10].value = 1
            # is empty
            try:
                row[11].value = 'True' if int(round(float(filament_amount))) <= EMPTY_THRESHOLD else 'False'
            except Exception:
                row[11].value = 'False'
            # optionally update roll weight if provided
            if roll_weight is not None:
                try:
                    row[9].value = int(round(float(roll_weight)))
                except Exception:
                    row[9].value = roll_weight
            updated = True
            break

    if updated:
        wb.save(EXCEL_PATH)
    return updated

def log_full_filament_data_web(brand, color, material, attr1, attr2, location, starting_weight, roll_weight):
    """
    Add a new filament roll row to the sheet. Generates a barcode and computes
    the initial filament amount as starting_weight - roll_weight.
    Returns a dict with the created entry.
    """
    wb, sheet = get_workbook_and_sheet()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    barcode = generate_barcode.generate_filament_barcode(brand, color, material, attr1, attr2, location, sheet)

    try:
        starting = float(starting_weight)
    except Exception:
        starting = 0.0
    try:
        roll_w = float(roll_weight)
    except Exception:
        roll_w = 0.0

    filament_amount = int(round(starting - roll_w))

    sheet.append([
        timestamp,
        barcode,
        brand,
        color,
        material,
        attr1,
        attr2,
        filament_amount,
        location,
        int(round(roll_w)),
        0,
        'False',
        'False'
    ])
    wb.save(EXCEL_PATH)

    return {
        "timestamp": timestamp,
        "barcode": barcode,
        "brand": brand,
        "color": color,
        "material": material,
        "attribute_1": attr1,
        "attribute_2": attr2,
        "filament_amount": filament_amount,
        "location": location,
        "roll_weight": int(round(roll_w))
    }