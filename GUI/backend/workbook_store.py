import os
import shutil
from datetime import datetime
from contextlib import contextmanager

import openpyxl

from backend.config import EXCEL_PATH

INVENTORY_SHEET = "Inventory"
EVENTS_SHEET = "UsageEvents"

INVENTORY_HEADERS = [
    "Timestamp",
    "Barcode",
    "Brand",
    "Color",
    "Material",
    "Attribute 1",
    "Attribute 2",
    "Filament Amount (g)",
    "Location",
    "Roll Weight (g)",
    "Times Logged Out",
    "Is Empty",
    "Is Favorite",
]

EVENT_HEADERS = [
    "Timestamp",
    "Event Type",
    "Barcode",
    "Brand",
    "Color",
    "Material",
    "Attribute 1",
    "Attribute 2",
    "Location",
    "Input Weight (g)",
    "Roll Weight (g)",
    "Filament Amount (g)",
    "Delta Used (g)",
    "Times Logged Out",
    "Source",
]

try:
    import portalocker
except Exception:
    portalocker = None


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip().lower() in ("1", "true", "yes", "on")
    return bool(value)


def _to_int(value, default):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _ensure_parent_dir():
    parent = os.path.dirname(EXCEL_PATH) or "."
    os.makedirs(parent, exist_ok=True)


def _load_backup_preferences():
    try:
        from backend import settings_store

        settings = settings_store.load_settings()
    except Exception:
        return False, 30

    enabled = _to_bool(settings.get("auto_backup_on_write"), False)
    retention_days = _to_int(settings.get("backup_retention_days"), 30)
    retention_days = max(1, min(retention_days, 3650))
    return enabled, retention_days


def _cleanup_old_backups(backup_dir, retention_days):
    cutoff = datetime.now().timestamp() - float(retention_days) * 86400.0
    try:
        for name in os.listdir(backup_dir):
            if not name.lower().endswith(".xlsx"):
                continue
            path = os.path.join(backup_dir, name)
            try:
                if os.path.getmtime(path) < cutoff:
                    os.remove(path)
            except Exception:
                continue
    except Exception:
        return


def _backup_workbook(retention_days):
    if not os.path.exists(EXCEL_PATH):
        return

    workbook_dir = os.path.dirname(EXCEL_PATH) or "."
    backup_dir = os.path.join(workbook_dir, "backups")
    os.makedirs(backup_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(EXCEL_PATH))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"{base_name}_{timestamp}.xlsx")

    try:
        shutil.copy2(EXCEL_PATH, backup_path)
    except Exception:
        return

    _cleanup_old_backups(backup_dir, retention_days)


def _ensure_headers(sheet, headers):
    if sheet.max_row < 1:
        sheet.append(headers)
        return

    row_one = [sheet.cell(row=1, column=idx + 1).value for idx in range(len(headers))]
    if all(value is None for value in row_one):
        for idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=idx).value = header


def _resolve_inventory_sheet(workbook, write=False):
    if INVENTORY_SHEET in workbook.sheetnames:
        sheet = workbook[INVENTORY_SHEET]
    elif workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]
        if write:
            sheet.title = INVENTORY_SHEET
    else:
        if not write:
            return None
        sheet = workbook.create_sheet(INVENTORY_SHEET, 0)

    if write and sheet is not None:
        _ensure_headers(sheet, INVENTORY_HEADERS)
    return sheet


def _resolve_events_sheet(workbook, write=False):
    if EVENTS_SHEET in workbook.sheetnames:
        sheet = workbook[EVENTS_SHEET]
        if write:
            _ensure_headers(sheet, EVENT_HEADERS)
        return sheet

    if not write:
        return None

    sheet = workbook.create_sheet(EVENTS_SHEET)
    _ensure_headers(sheet, EVENT_HEADERS)
    return sheet


def _create_workbook():
    workbook = openpyxl.Workbook()
    inventory = workbook.active
    inventory.title = INVENTORY_SHEET
    inventory.append(INVENTORY_HEADERS)

    events = workbook.create_sheet(EVENTS_SHEET)
    events.append(EVENT_HEADERS)

    workbook.save(EXCEL_PATH)
    workbook.close()


@contextmanager
def _workbook_lock(write=False):
    if portalocker is None:
        yield
        return

    lock_path = EXCEL_PATH + ".lock"
    handle = open(lock_path, "a+")
    lock_flag = portalocker.LOCK_EX if write else portalocker.LOCK_SH
    portalocker.lock(handle, lock_flag)
    try:
        yield
    finally:
        try:
            portalocker.unlock(handle)
        finally:
            handle.close()


@contextmanager
def open_workbook(write=False):
    _ensure_parent_dir()
    with _workbook_lock(write=write):
        if not os.path.exists(EXCEL_PATH):
            _create_workbook()

        workbook = openpyxl.load_workbook(EXCEL_PATH)
        inventory = _resolve_inventory_sheet(workbook, write=write)
        events = _resolve_events_sheet(workbook, write=write)

        try:
            yield workbook, inventory, events
            if write:
                backup_enabled, backup_retention_days = _load_backup_preferences()
                if backup_enabled:
                    _backup_workbook(backup_retention_days)
                workbook.save(EXCEL_PATH)
        finally:
            workbook.close()


def list_inventory_rows():
    with open_workbook(write=False) as (_, inventory, _):
        if inventory is None:
            return []
        return [row for row in inventory.iter_rows(min_row=2, values_only=True)]
