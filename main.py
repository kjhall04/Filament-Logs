import openpyxl
import time
import log_modules
import generate_barcode

# File path for the Excel workbook
FILE_PATH = 'filament_inventory.xlsx'

# Initialize workbook and sheet
try:
    workbook = openpyxl.load_workbook(FILE_PATH)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Timestamp', 'Barcode', 'Brand', 'Color', 'Material', 'Attribute 1', 'Attribute 2', 'Filament Amount (g)', 'Location', 'Roll Weight (g)'])  # Add headers

def get_roll_weight(barcode: str, sheet) -> float:
    """Retrieve the roll weight from the spreadsheet based on the barcode."""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == barcode:  # Barcode is in the second column
            return row[-1]    # Roll Weight (g) is the last column
    raise ValueError(f"Roll weight not found for barcode: {barcode}")

def log_filament_data(generate):
    """
    Main loop to log filament data, including barcode, weight, and decoded filament name.
    Saves the data to an Excel file and handles user interruptions gracefully.
    """
    print("Press Ctrl+C to exit the program.")

    if not generate:
        while True:
            try:
                # Gather data
                barcode = log_modules.get_barcode()
                brand, color, material, attribute_1, attribute_2, location = log_modules.decode_barcode(barcode)

                # Display log info
                print(f'Logging weight for filament: {brand} {color} {material} {attribute_1} {attribute_2}')

                roll_weight = get_roll_weight(barcode, sheet)
                filament_amount = log_modules.get_current_weight(roll_weight)

                # Append data to the sheet
                timestamp = time.strftime('%Y-%m-%d %H:%M:%S')

                sheet.append([timestamp, barcode, brand, color, material, attribute_1, attribute_2, filament_amount, location])
                workbook.save(FILE_PATH)

                print(f'''Logged: 
                      Timestamp: {timestamp}, 
                      Barcode: {barcode}, 
                      Brand: {brand}, 
                      Color: {color}, 
                      Material: {material}, 
                      Attribute 1: {attribute_1}, 
                      Attribute 2: {attribute_2}, 
                      Filament Amount: {filament_amount}, 
                      Location: {location}''')

            except KeyboardInterrupt:
                print('\nExiting program. Goodbye!')
                break
            except Exception as e:
                print(f"An error occurred: {e}")

    else:
        while True:
            try:
                brand = input('Enter the brand: ')
                color = input('Enter the color: ')
                material = input('Enter the material: ')
                attribute_1 = input('Enter the first attribute: ')
                attribute_2 = input('Enter the second attribute: ')
                location = input('Enter the location of the filament: ')

                barcode = generate_barcode.generate_filament_barcode(brand, color, material, attribute_1, attribute_2, location, sheet)
                brand, color, material, attribute_1, attribute_2, location = log_modules.decode_barcode(barcode)
                print(f'New barcode for {brand} {color} {material} {attribute_1} {attribute_2} in {location} is {barcode}')
                # Append data to the sheet
                timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
                roll_weight, filament_amount = log_modules.get_starting_weight()

                sheet.append([timestamp, barcode, brand, color, material, attribute_1, attribute_2, filament_amount, location, roll_weight])
                workbook.save(FILE_PATH)

                print(f'''Logged: 
                      Timestamp: {timestamp}, 
                      Barcode: {barcode}, 
                      Brand: {brand}, 
                      Color: {color}, 
                      Material: {material}, 
                      Attribute 1: {attribute_1}, 
                      Attribute 2: {attribute_2}, 
                      Filament Amount: {filament_amount}, 
                      Location: {location}, 
                      Roll Weight: {roll_weight}''')
                
            except KeyboardInterrupt:
                print('\nExiting program. Goodbye!')
                break
            except ValueError as e:
                print(e)

if __name__ == '__main__':
    log_filament_data(generate=False)